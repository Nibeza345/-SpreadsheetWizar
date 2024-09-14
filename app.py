import openpyxl as xl
from openpyxl.chart import BarChart, Reference
import logging
import shutil
import os

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def process_workbook(filename, sheet_name=None, adjustment_factor=0.9, create_backup=True):
    """
    Processes an Excel workbook to adjust prices in a specific column
    and adds a bar chart based on the corrected prices.

    Args:
    - filename (str): The path to the Excel file.
    - sheet_name (str, optional): The name of the sheet to process (default is None, which will use the first sheet).
    - adjustment_factor (float, optional): The factor by which to adjust prices (default is 0.9).
    - create_backup (bool, optional): Whether to create a backup of the workbook before saving (default is True).
    """
    try:
   
        wb = xl.load_workbook(filename)
        
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                logging.error(f"Sheet '{sheet_name}' not found in the workbook.")
                return
            sheet = wb[sheet_name]
        else:
            sheet = wb.active
            logging.info(f"No sheet name provided, using the first sheet: '{sheet.title}'")
        
        
        if create_backup:
            backup_filename = f"{filename}.backup"
            shutil.copyfile(filename, backup_filename)
            logging.info(f"Backup created at '{backup_filename}'")
        
    
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row, 3)  
            if isinstance(cell.value, (int, float)):
                corrected_price = cell.value * adjustment_factor  
                corrected_price_cell = sheet.cell(row, 4)  
                corrected_price_cell.value = corrected_price
            else:
                logging.warning(f"Non-numeric value in row {row}, column 3.")
        
        values = Reference(
            sheet,
            min_row=2,
            max_row=sheet.max_row,
            min_col=4,  
            max_col=4
        )

        if sheet.max_row > 1:
            chart = BarChart()
            chart.add_data(values, titles_from_data=False)
            chart.title = "Corrected Prices Chart"
            chart.x_axis.title = "Product"
            chart.y_axis.title = "Corrected Price"
            sheet.add_chart(chart, 'G2')  
            logging.info(f"Chart added to the sheet '{sheet.title}'.")
        else:
            logging.warning("No data to create a chart.")
        
        wb.save(filename)
        logging.info(f"Workbook '{filename}' processed successfully and saved.")
    
    except FileNotFoundError:
        logging.error(f"Error: File '{filename}' not found.")
    except Exception as e:
        logging.error(f"An error occurred: {e}")


process_workbook('transactions.xlsx')
